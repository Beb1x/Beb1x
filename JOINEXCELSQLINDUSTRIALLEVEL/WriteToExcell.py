import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import configparser

def Excell_update(all_results):
    config = configparser.ConfigParser()
    config.read('C:\\Serializare\\config\\parametri.config')
    excel_folder_path = config.get('Settings', 'PathExcell')

    for filename in os.listdir(excel_folder_path):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            excel_file_path = os.path.join(excel_folder_path, filename)

            # Load the Excel file
            wb = load_workbook(excel_file_path)

            # Check if the "Sheet1" sheet exists in the Excel file
            if 'Sheet1' in wb.sheetnames:
                sheet = wb['Sheet1']

                # Find the column index of "raspuns" in the first row
                raspuns_column_index = None
                for cell in sheet[1]:
                    if cell.value == 'raspuns':
                        raspuns_column_index = cell.column
                        break

                if raspuns_column_index is not None:
                    # Update the "raspuns" column with the specified values
                    for i, result in enumerate(all_results, start=2):  # Assuming data starts from row 2
                        sheet.cell(row=i, column=raspuns_column_index, value=result)

                    # Save the updated Excel file
                    wb.save(excel_file_path)
                    print(f'Successfully updated "raspuns" column in {filename}')
                else:
                    # If "raspuns" column doesn't exist, create it
                    column_letter = get_column_letter(sheet.max_column + 1)
                    sheet[column_letter + '1'] = 'raspuns'

                    # Update the "raspuns" column with the specified values
                    for i, result in enumerate(all_results, start=2):  # Assuming data starts from row 2
                        sheet.cell(row=i, column=sheet.max_column, value=result)

                    # Save the updated Excel file
                    wb.save(excel_file_path)
                    print(f'Created and successfully updated "raspuns" column in {filename}')
            else:
                # If "Sheet1" doesn't exist, create it and add "raspuns" column
                sheet = wb.create_sheet('Sheet1')
                sheet['A1'] = 'raspuns'

                # Update the "raspuns" column with the specified values
                for i, result in enumerate(all_results, start=2):  # Assuming data starts from row 2
                    sheet.cell(row=i, column=1, value=result)

                # Save the updated Excel file
                wb.save(excel_file_path)
                print(f'Created and successfully updated "raspuns" column in {filename}')



#all_results = ['New Value 1', 'New Value 2']
#Excell_update(all_results)